"""
Michael Credit Card processor - categorizes expenses using AI
"""
import pandas as pd
from io import BytesIO
from typing import Dict, List, Any, Optional
import json
import os

from openai import OpenAI
from dotenv import load_dotenv

load_dotenv()
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# Mesmas categorias do categorizer.py
EXPENSE_CATEGORIES = [
    "Airfare",
    "Computer Equipment",
    "Travel - Event",
    "Gifts",
    "Lodging",
    "Miscellaneous",
    "Office Supplies",
    "IT Subscriptions",
    "Training",
    "Brazil Insurance",
    "Personal Expenses",
    "Membership Dues",
    "Printing",
    "Rippling Wire Deduction",
    "Ground Transportation",
    "Meals & Entertainment",
    "Conferences & Seminars",
    "Telephone/Internet",
    "Wellhub Reimbursement",
    "Pantry Food",
    "Travel Agent Fees",
    "Delivery and Postage",
    "Venue - Event",
    "Catering - Event",
    "Printing - Event",
    "Tech/AV - Event",
    "Other - Event",
    "Board meetings",
    "Due Diligence - Portfolio Company",
    "Due Diligence - New Deals",
]

# Mapeamento de categorias do Amex para nossas categorias
AMEX_CATEGORY_MAPPING = {
    "Travel-Airline": "Airfare",
    "Travel-Lodging": "Lodging",
    "Travel-Travel Agencies": "Travel Agent Fees",
    "Transportation-Taxis & Coach": "Ground Transportation",
    "Restaurant-Restaurant": "Meals & Entertainment",
    "Communications-Telephone Comm": "Telephone/Internet",
    "Merchandise & Supplies-Computer Supplies": "Computer Equipment",
    "Merchandise & Supplies-General Retail": "Miscellaneous",
    "Merchandise & Supplies-Mail Order": "Delivery and Postage",
    "Business Services-Professional Services": "Miscellaneous",
    "Business Services-Other Services": "Miscellaneous",
    "Other-Miscellaneous": "Miscellaneous",
}

# Keywords para regras baseadas em Extended Details
AIRLINE_KEYWORDS = ["AIRLINE", "AIRWAYS", "AMERICAN AIR", "DELTA AIR", "UNITED AIR", "LATAM", "GOL LINHAS", "GOL AIR", "AZUL", "JETBLUE", "SOUTHWEST", "PASSENGER TICKET", "FLIGHT", "AIR CANADA", "BRITISH AIR", "LUFTHANSA", "EMIRATES", "QATAR", "TAM ", "TAP ", "AVIANCA"]
HOTEL_KEYWORDS = ["HOTEL", "MARRIOTT", "HILTON", "HYATT", "IHG", "AIRBNB", "VRBO", "LODGING", "MELIA", "RESORT", "SUITES", "INN ", " INN", "MOTEL", "HOSTEL", "POUSADA", "SHERATON", "WESTIN", "INTERCONTINENTAL", "HOLIDAY INN", "BEST WESTERN", "RADISSON", "COSTA BAVARO"]
RESTAURANT_KEYWORDS = ["RESTAURANT", "CAFE", "COFFEE", "STARBUCKS", "BURGER", "PIZZA", "GRILL", "BAR ", " BAR", "BISTRO", "DINER", "FOOD", "EATERY", "BAKERY", "PADARIA", "LANCHONETE", "CHURRASCARIA", "STEAKHOUSE", "SUSHI", "JAPANESE", "ITALIAN", "CHINESE", "MEXICAN", "THAI"]
GROUND_TRANSPORT_KEYWORDS = ["UBER", "LYFT", "TAXI", "CAB ", "LIMO", "CABIFY", "99 ", "TAXIS", "99APP", "YELLOW CAB", "SHUTTLE", "TRANSFER", "CAR SERVICE", "BLACK CAR"]
IT_KEYWORDS = ["AMAZON WEB", "AWS", "MICROSOFT", "GOOGLE CLOUD", "ZOOM", "SLACK", "NOTION", "GITHUB", "ADOBE", "DROPBOX", "OPENAI", "CHATGPT", "STRIPE", "TWILIO", "HEROKU", "DIGITAL OCEAN"]
TRAVEL_AGENCY_KEYWORDS = ["EXPEDIA", "BOOKING.COM", "HOTELS.COM", "KAYAK", "TRAVEL AGENCY", "PRICELINE", "ORBITZ", "TRAVELOCITY", "DESPEGAR", "DECOLAR", "CVC VIAGENS"]
TELECOM_KEYWORDS = ["AT&T", "VERIZON", "T-MOBILE", "SPRINT", "CLARO", "VIVO", "TIM ", "OI ", "TELEFONICA", "COMCAST", "SPECTRUM", "TELECOM"]


def rule_based_category(extended_details: str, amex_category: str) -> Optional[str]:
    """
    Aplica regras determinísticas baseadas em Extended Details e Category do Amex
    """
    details_upper = (extended_details or "").upper()
    amex_cat = (amex_category or "").strip()
    
    # 1. Regras baseadas em Extended Details (prioridade máxima)
    if any(kw in details_upper for kw in AIRLINE_KEYWORDS):
        return "Airfare"
    
    if any(kw in details_upper for kw in HOTEL_KEYWORDS):
        return "Lodging"
    
    if any(kw in details_upper for kw in TRAVEL_AGENCY_KEYWORDS):
        return "Travel Agent Fees"
    
    if any(kw in details_upper for kw in RESTAURANT_KEYWORDS):
        return "Meals & Entertainment"
    
    if any(kw in details_upper for kw in GROUND_TRANSPORT_KEYWORDS):
        return "Ground Transportation"
    
    if any(kw in details_upper for kw in IT_KEYWORDS):
        return "IT Subscriptions"
    
    if any(kw in details_upper for kw in TELECOM_KEYWORDS):
        return "Telephone/Internet"
    
    # 2. Fallback para categoria do Amex (se disponível)
    if amex_cat and amex_cat in AMEX_CATEGORY_MAPPING:
        return AMEX_CATEGORY_MAPPING[amex_cat]
    
    # 3. Se não encontrou nada, retorna None para ir para AI
    return None


def normalize_category(raw: str) -> str:
    """Normaliza categoria retornada pelo modelo"""
    if not raw:
        return ""
    normalized = raw.strip()
    for valid in EXPENSE_CATEGORIES:
        if normalized.lower() == valid.lower():
            return valid
    return ""


def build_llm_prompt(batch_data: List[Dict]) -> str:
    """Constrói prompt para o LLM categorizar"""
    categories_list = "\n".join(f"- {cat}" for cat in EXPENSE_CATEGORIES)
    
    items_text = "\n".join([
        f"{i+1}. Extended Details: \"{item['extended_details']}\" | Amex Category: \"{item['amex_category']}\""
        for i, item in enumerate(batch_data)
    ])
    
    prompt = f"""You are an expense categorization assistant. Categorize each expense into one of these categories:

{categories_list}

RULES:
1. "Travel-Airline" or anything with airlines -> "Airfare"
2. "Travel-Lodging" or hotels -> "Lodging"
3. "Restaurant-Restaurant" or food places -> "Meals & Entertainment"
4. "Transportation-Taxis & Coach" or Uber/Lyft/Taxi -> "Ground Transportation"
5. "Communications-Telephone Comm" -> "Telephone/Internet"
6. "Travel-Travel Agencies" -> "Travel Agent Fees"
7. "Merchandise & Supplies-Computer Supplies" -> "Computer Equipment"
8. If you can't determine confidently, use "Miscellaneous"

Here are {len(batch_data)} expenses to categorize:
{items_text}

Return ONLY a valid JSON array with {len(batch_data)} category strings.
Example: ["Airfare", "Lodging", "Meals & Entertainment"]
"""
    return prompt


def categorize_with_llm(items: List[Dict]) -> List[str]:
    """Chama o LLM para categorizar itens"""
    if not items:
        return []
    
    batch_size = 25
    all_categories = []
    
    for i in range(0, len(items), batch_size):
        batch = items[i:i + batch_size]
        prompt = build_llm_prompt(batch)
        
        try:
            response = client.chat.completions.create(
                model="gpt-4.1",
                messages=[
                    {"role": "system", "content": "You are an expense categorization assistant. Respond ONLY with valid JSON arrays."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_tokens=2000
            )
            
            result_text = response.choices[0].message.content.strip()
            batch_categories = json.loads(result_text)
            
            if not isinstance(batch_categories, list):
                raise ValueError("Response is not a JSON array")
            
            # Ajusta tamanho se necessário
            if len(batch_categories) != len(batch):
                if len(batch_categories) > len(batch):
                    batch_categories = batch_categories[:len(batch)]
                else:
                    batch_categories += ["Miscellaneous"] * (len(batch) - len(batch_categories))
                    
        except Exception as e:
            print(f"[ERROR] LLM categorization failed: {e}")
            batch_categories = ["Miscellaneous"] * len(batch)
        
        all_categories.extend(batch_categories)
        print(f"Categorized batch {i // batch_size + 1}: {len(batch_categories)} items")
    
    return all_categories


def process_michael_file(file_content: bytes, filename: str) -> Dict[str, Any]:
    """
    Processa arquivo Excel do Michael e retorna transações com categorias originais.
    Se a categoria do arquivo for uma categoria válida do sistema, usa diretamente.
    """
    # Ler arquivo
    if filename.endswith('.csv'):
        df = pd.read_csv(BytesIO(file_content))
    else:
        df = pd.read_excel(BytesIO(file_content))
    
    # Debug: mostrar colunas disponíveis
    print(f"[DEBUG] Colunas do arquivo: {df.columns.tolist()}")
    print(f"[DEBUG] Primeira linha: {df.iloc[0].to_dict() if len(df) > 0 else 'vazio'}")
    
    # Set de categorias válidas para verificação rápida
    valid_categories_set = set(EXPENSE_CATEGORIES)
    
    # Remover linhas de cabeçalho duplicado (se existir coluna Date)
    if 'Date' in df.columns:
        df = df[df['Date'] != 'Date']
        df = df.dropna(subset=['Date'])
    
    # Resetar índice para garantir IDs únicos
    df = df.reset_index(drop=True)
    
    # Converter para lista de transações
    transactions = []
    for idx, row in df.iterrows():
        ext_details = str(row.get('Extended Details', '')) if pd.notna(row.get('Extended Details')) else ''
        file_category = str(row.get('Category', '')) if pd.notna(row.get('Category')) else ''
        
        # Verificar se a categoria do arquivo é uma categoria válida do sistema
        # Se for (ex: "Airfare", "Lodging"), usa como ai_category
        # Se não for (ex: "Travel-Airline"), guarda como amex_category para processar via AI
        ai_category = ""
        amex_category = ""
        
        if file_category in valid_categories_set:
            # Categoria já é válida - usar diretamente
            ai_category = file_category
            print(f"[DEBUG] TX {idx+1}: Category '{file_category}' is valid, using as ai_category")
        else:
            # Categoria do Amex original - precisa ser mapeada/processada
            amex_category = file_category
            if idx < 3:
                print(f"[DEBUG] TX {idx+1}: Category '{file_category}' needs mapping, stored as amex_category")
        
        tx = {
            "id": idx + 1,
            "date": str(row.get('Date', '')),
            "description": str(row.get('Description', '')),
            "notes": str(row.get('Unnamed: 0', row.get('Notes', ''))),
            "amount": float(row.get('Amount', 0)) if pd.notna(row.get('Amount')) else 0,
            "extended_details": ext_details,
            "amex_category": amex_category,
            "city_state": str(row.get('City/State', '')) if pd.notna(row.get('City/State')) else '',
            "ai_category": ai_category
        }
        
        transactions.append(tx)
    
    already_categorized = sum(1 for t in transactions if t["ai_category"])
    print(f"[DEBUG] Total: {len(transactions)} transactions, {already_categorized} already categorized from file")
    
    return {
        "transactions": transactions,
        "total_transactions": len(transactions),
        "total_amount": sum(t["amount"] for t in transactions),
        "already_categorized": already_categorized
    }


def categorize_michael_transactions(transactions: List[Dict]) -> List[Dict]:
    """
    Categoriza transações usando regras + AI.
    Mantém categorias que já vieram preenchidas do arquivo.
    """
    if not transactions:
        return transactions
    
    uncategorized_indices = []
    
    # 1ª passada: regras determinísticas (apenas para os que não têm ai_category)
    for i, tx in enumerate(transactions):
        # Se já tem ai_category válida (veio do arquivo), mantém
        if tx.get("ai_category") and tx["ai_category"].strip():
            print(f"[DEBUG] TX {i+1}: Keeping existing ai_category: '{tx['ai_category']}'")
            continue
            
        # Tenta categorizar com regras
        rule_cat = rule_based_category(tx.get("extended_details", ""), tx.get("amex_category", ""))
        if rule_cat:
            tx["ai_category"] = rule_cat
        else:
            tx["ai_category"] = ""
            uncategorized_indices.append(i)
    
    already_done = len(transactions) - len([t for t in transactions if not t.get("ai_category")])
    print(f"Rule-based: {already_done} categorized (including from file), {len(uncategorized_indices)} need AI")
    
    if not uncategorized_indices:
        return transactions
    
    # 2ª passada: LLM para os não categorizados
    items_for_llm = [
        {
            "extended_details": transactions[i].get("extended_details", ""),
            "amex_category": transactions[i].get("amex_category", "")
        }
        for i in uncategorized_indices
    ]
    
    llm_categories = categorize_with_llm(items_for_llm)
    
    # 3ª passada: aplica categorias do LLM
    for idx, tx_index in enumerate(uncategorized_indices):
        raw_cat = llm_categories[idx] if idx < len(llm_categories) else "Miscellaneous"
        normalized = normalize_category(raw_cat)
        transactions[tx_index]["ai_category"] = normalized or "Miscellaneous"
    
    return transactions


def export_michael_to_excel(transactions: List[Dict]) -> bytes:
    """
    Exporta transações categorizadas para Excel
    """
    # Criar DataFrame
    df = pd.DataFrame(transactions)
    
    # Reordenar e renomear colunas
    columns_order = ['date', 'description', 'notes', 'amount', 'extended_details', 'amex_category', 'ai_category', 'city_state']
    columns_rename = {
        'date': 'Date',
        'description': 'Description',
        'notes': 'Notes',
        'amount': 'Amount',
        'extended_details': 'Extended Details',
        'amex_category': 'Original Category',
        'ai_category': 'AI Category',
        'city_state': 'City/State'
    }
    
    df = df[[c for c in columns_order if c in df.columns]]
    df = df.rename(columns=columns_rename)
    
    # Criar Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Categorized Expenses')
        
        # Formatar
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        worksheet = writer.sheets['Categorized Expenses']
        
        # Ajustar largura das colunas
        for idx, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            ) + 2
            worksheet.column_dimensions[get_column_letter(idx)].width = min(max_length, 40)
        
        # Header style
        header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Number format for Amount column
        amount_col = None
        for idx, col in enumerate(df.columns, 1):
            if col == 'Amount':
                amount_col = idx
                break
        
        if amount_col:
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=amount_col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
    
    output.seek(0)
    return output.getvalue()


# =====================================================
# BigQuery Database Operations
# =====================================================
from google.cloud import bigquery
from datetime import datetime
import uuid

PROJECT_ID = "automatic-bond-462415-h6"
DATASET_ID = "finance"
TABLE_ID = "michael_expenses"
VALOR_TABLE_ID = "valor_expenses"

# Lazy initialization of BigQuery client
_bq_client = None

def get_bq_client():
    """Get or create BigQuery client with lazy initialization"""
    global _bq_client
    if _bq_client is None:
        _bq_client = bigquery.Client(project=PROJECT_ID)
    return _bq_client


def get_michael_expenses(year: int = None, limit: int = 1000) -> List[Dict[str, Any]]:
    """Get Michael expenses from the database."""
    
    if year is None:
        year = datetime.now().year
    
    query = f"""
    SELECT 
        id,
        date,
        description,
        card_member,
        amount,
        category,
        project,
        batch_id,
        created_at,
        synced_to_valor,
        valor_expense_id
    FROM `{PROJECT_ID}.{DATASET_ID}.{TABLE_ID}`
    WHERE EXTRACT(YEAR FROM date) = {year}
    ORDER BY date DESC
    LIMIT {limit}
    """
    
    expenses = []
    try:
        results = get_bq_client().query(query).result()
        for row in results:
            expenses.append({
                "id": row.id,
                "date": row.date.isoformat() if row.date else None,
                "description": row.description or "",
                "card_member": row.card_member or "",
                "amount": float(row.amount) if row.amount else 0,
                "category": row.category or "",
                "project": row.project or "",
                "batch_id": row.batch_id or "",
                "created_at": row.created_at.isoformat() if row.created_at else None,
                "synced_to_valor": row.synced_to_valor or False,
                "valor_expense_id": row.valor_expense_id or ""
            })
    except Exception as e:
        print(f"Error querying michael_expenses: {e}")
        raise e
    
    return expenses


def get_michael_batches() -> List[Dict[str, Any]]:
    """Get all batches of Michael expenses."""
    
    query = f"""
    SELECT 
        batch_id,
        MIN(created_at) as created_at,
        COUNT(*) as transaction_count,
        SUM(amount) as total_amount
    FROM `{PROJECT_ID}.{DATASET_ID}.{TABLE_ID}`
    WHERE batch_id IS NOT NULL
    GROUP BY batch_id
    ORDER BY created_at DESC
    """
    
    batches = []
    try:
        results = get_bq_client().query(query).result()
        for row in results:
            batches.append({
                "batch_id": row.batch_id,
                "created_at": row.created_at.isoformat() if row.created_at else None,
                "transaction_count": row.transaction_count,
                "total_amount": float(row.total_amount) if row.total_amount else 0
            })
    except Exception as e:
        print(f"Error querying michael batches: {e}")
        raise e
    
    return batches


def add_michael_expenses_to_db(expenses: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Add a batch of Michael expenses to the database using INSERT (not streaming)."""
    
    batch_id = f"michael_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    if not expenses:
        return {"success": True, "batch_id": batch_id, "added_count": 0, "total_amount": 0}
    
    # Build INSERT statement (not streaming - allows immediate UPDATE)
    values_parts = []
    total_amount = 0
    today = datetime.now().strftime('%Y-%m-%d')  # Default date if none provided
    
    for exp in expenses:
        row_id = str(uuid.uuid4())
        
        # Parse date - use today as fallback
        date_str = exp.get("date", "") or ""
        date_str = date_str.strip()
        date_sql = f"DATE'{today}'"  # Default to today
        
        if date_str:
            parsed = False
            try:
                # Remove time portion if present (handles "2024-12-31 00:00:00" format)
                date_only = date_str.split(" ")[0].split("T")[0]
                
                # Try different date formats
                for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"]:
                    try:
                        date_obj = datetime.strptime(date_only, fmt)
                        date_sql = f"DATE'{date_obj.strftime('%Y-%m-%d')}'"
                        parsed = True
                        break
                    except:
                        continue
                
                # Handle invalid dates like 2024-13-31 (month > 12)
                if not parsed and "-" in date_only:
                    parts = date_only.split("-")
                    if len(parts) == 3:
                        year, month, day = parts[0], parts[1], parts[2]
                        # Fix invalid month (>12) - use December
                        if month.isdigit() and int(month) > 12:
                            month = "12"
                        # Fix invalid day (>31) - use last day
                        if day.isdigit() and int(day) > 31:
                            day = "31"
                        date_sql = f"DATE'{year}-{month.zfill(2)}-{day.zfill(2)}'"
                        print(f"[DEBUG] Fixed invalid date {date_str} -> {year}-{month}-{day}")
            except Exception as e:
                print(f"[DEBUG] Date parse error for '{date_str}': {e}")
                pass  # Keep default (today)
        
        # Escape strings for SQL
        desc = (exp.get("description", "") or "").replace("\\", "\\\\").replace("'", "\\'")
        card_member = (exp.get("card_member", "Michael Nicklas") or "Michael Nicklas").replace("\\", "\\\\").replace("'", "\\'")
        category = (exp.get("category", "") or "").replace("\\", "\\\\").replace("'", "\\'")
        project = (exp.get("project", "") or "").replace("\\", "\\\\").replace("'", "\\'")
        amount = float(exp.get("amount", 0) or 0)
        total_amount += amount
        
        values_parts.append(f"""
            ('{row_id}', {date_sql}, '{desc}', '{card_member}', {amount}, '{category}', '{project}', '{batch_id}', CURRENT_TIMESTAMP(), FALSE, NULL)
        """)
    
    # Insert in batches to avoid query size limits
    table_ref = f"{PROJECT_ID}.{DATASET_ID}.{TABLE_ID}"
    batch_size = 50
    
    try:
        for i in range(0, len(values_parts), batch_size):
            batch = values_parts[i:i+batch_size]
            insert_query = f"""
            INSERT INTO `{table_ref}`
            (id, date, description, card_member, amount, category, project, batch_id, created_at, synced_to_valor, valor_expense_id)
            VALUES {','.join(batch)}
            """
            get_bq_client().query(insert_query).result()
            
    except Exception as e:
        print(f"Error adding michael expenses: {e}")
        raise e
    
    return {
        "success": True,
        "batch_id": batch_id,
        "added_count": len(values_parts),
        "total_amount": total_amount
    }


def update_michael_expense(expense_id: str, updates: Dict[str, Any]) -> Dict[str, Any]:
    """Update a Michael expense.
    
    Note: May fail if data is in BigQuery streaming buffer (first 90 min after insert).
    """
    
    # Build SET clause
    set_parts = []
    params = []
    
    if "category" in updates:
        set_parts.append("category = @category")
        params.append(bigquery.ScalarQueryParameter("category", "STRING", updates["category"]))
    
    if "project" in updates:
        set_parts.append("project = @project")
        params.append(bigquery.ScalarQueryParameter("project", "STRING", updates["project"]))
    
    if "description" in updates:
        set_parts.append("description = @description")
        params.append(bigquery.ScalarQueryParameter("description", "STRING", updates["description"]))
    
    if "amount" in updates:
        set_parts.append("amount = @amount")
        params.append(bigquery.ScalarQueryParameter("amount", "FLOAT64", float(updates["amount"])))
    
    if "date" in updates:
        set_parts.append("date = @date")
        params.append(bigquery.ScalarQueryParameter("date", "DATE", updates["date"]))
    
    if not set_parts:
        return {"success": True, "synced_valor": False}
    
    params.append(bigquery.ScalarQueryParameter("id", "STRING", expense_id))
    
    query = f"""
    UPDATE `{PROJECT_ID}.{DATASET_ID}.{TABLE_ID}`
    SET {", ".join(set_parts)}
    WHERE id = @id
    """
    
    job_config = bigquery.QueryJobConfig(query_parameters=params)
    
    try:
        get_bq_client().query(query, job_config=job_config).result()
    except Exception as e:
        error_msg = str(e).lower()
        if "streaming buffer" in error_msg:
            raise Exception("Cannot update: Data is in streaming buffer. Please wait ~90 minutes after upload before editing.")
        print(f"Error updating michael expense: {e}")
        raise e
    
    # Sync to valor_expenses if already synced
    synced_valor = False
    try:
        check_query = f"""
        SELECT valor_expense_id, synced_to_valor 
        FROM `{PROJECT_ID}.{DATASET_ID}.{TABLE_ID}`
        WHERE id = @id
        """
        check_config = bigquery.QueryJobConfig(
            query_parameters=[bigquery.ScalarQueryParameter("id", "STRING", expense_id)]
        )
        result = list(get_bq_client().query(check_query, job_config=check_config).result())
        
        if result and result[0].synced_to_valor and result[0].valor_expense_id:
            valor_set_parts = []
            valor_params = []
            
            if "category" in updates:
                valor_set_parts.append("category = @category")
                valor_params.append(bigquery.ScalarQueryParameter("category", "STRING", updates["category"]))
            
            if "project" in updates:
                valor_set_parts.append("project = @project")
                valor_params.append(bigquery.ScalarQueryParameter("project", "STRING", updates["project"]))
            
            if "date" in updates:
                valor_set_parts.append("date = @date")
                valor_set_parts.append("year = EXTRACT(YEAR FROM @date)")
                valor_set_parts.append("month = EXTRACT(MONTH FROM @date)")
                valor_params.append(bigquery.ScalarQueryParameter("date", "DATE", updates["date"]))
            
            if valor_set_parts:
                valor_params.append(bigquery.ScalarQueryParameter("valor_id", "STRING", result[0].valor_expense_id))
                valor_query = f"""
                UPDATE `{PROJECT_ID}.{DATASET_ID}.{VALOR_TABLE_ID}`
                SET {", ".join(valor_set_parts)}
                WHERE id = @valor_id
                """
                valor_config = bigquery.QueryJobConfig(query_parameters=valor_params)
                get_bq_client().query(valor_query, job_config=valor_config).result()
                synced_valor = True
    except Exception as e:
        print(f"Error syncing to valor: {e}")
    
    return {"success": True, "synced_valor": synced_valor}


def delete_michael_expense(expense_id: str) -> Dict[str, Any]:
    """Delete a single Michael expense."""
    
    query = f"""
    SELECT amount, valor_expense_id
    FROM `{PROJECT_ID}.{DATASET_ID}.{TABLE_ID}`
    WHERE id = @id
    """
    job_config = bigquery.QueryJobConfig(
        query_parameters=[bigquery.ScalarQueryParameter("id", "STRING", expense_id)]
    )
    
    try:
        result = list(get_bq_client().query(query, job_config=job_config).result())
        if not result:
            raise Exception("Expense not found")
        
        amount = float(result[0].amount) if result[0].amount else 0
        valor_expense_id = result[0].valor_expense_id
        
        delete_query = f"""
        DELETE FROM `{PROJECT_ID}.{DATASET_ID}.{TABLE_ID}`
        WHERE id = @id
        """
        get_bq_client().query(delete_query, job_config=job_config).result()
        
        if valor_expense_id:
            valor_delete_query = f"""
            DELETE FROM `{PROJECT_ID}.{DATASET_ID}.{VALOR_TABLE_ID}`
            WHERE id = @valor_id
            """
            valor_config = bigquery.QueryJobConfig(
                query_parameters=[bigquery.ScalarQueryParameter("valor_id", "STRING", valor_expense_id)]
            )
            get_bq_client().query(valor_delete_query, job_config=valor_config).result()
        
        return {"success": True, "deleted_amount": amount}
        
    except Exception as e:
        print(f"Error deleting michael expense: {e}")
        raise e


def delete_michael_batch(batch_id: str) -> Dict[str, Any]:
    """Delete an entire batch of Michael expenses."""
    
    query = f"""
    SELECT COUNT(*) as count, SUM(amount) as total, 
           ARRAY_AGG(valor_expense_id IGNORE NULLS) as valor_ids
    FROM `{PROJECT_ID}.{DATASET_ID}.{TABLE_ID}`
    WHERE batch_id = @batch_id
    """
    job_config = bigquery.QueryJobConfig(
        query_parameters=[bigquery.ScalarQueryParameter("batch_id", "STRING", batch_id)]
    )
    
    try:
        result = list(get_bq_client().query(query, job_config=job_config).result())
        if not result or result[0].count == 0:
            raise Exception("Batch not found")
        
        count = result[0].count
        total = float(result[0].total) if result[0].total else 0
        valor_ids = result[0].valor_ids or []
        
        delete_query = f"""
        DELETE FROM `{PROJECT_ID}.{DATASET_ID}.{TABLE_ID}`
        WHERE batch_id = @batch_id
        """
        get_bq_client().query(delete_query, job_config=job_config).result()
        
        if valor_ids:
            valor_ids_str = ", ".join([f"'{vid}'" for vid in valor_ids if vid])
            if valor_ids_str:
                valor_delete_query = f"""
                DELETE FROM `{PROJECT_ID}.{DATASET_ID}.{VALOR_TABLE_ID}`
                WHERE id IN ({valor_ids_str})
                """
                get_bq_client().query(valor_delete_query).result()
        
        return {
            "success": True,
            "deleted_count": count,
            "total_amount": total
        }
        
    except Exception as e:
        print(f"Error deleting michael batch: {e}")
        raise e


def sync_michael_to_valor(expense_ids: List[str] = None) -> Dict[str, Any]:
    """Sync Michael expenses to valor_expenses using MERGE (upsert).
    
    Uses MERGE to prevent duplicates - if michael_<id> exists, UPDATE; otherwise INSERT.
    This approach avoids streaming buffer issues with UPDATE.
    """
    
    if expense_ids:
        ids_str = ", ".join([f"'{id}'" for id in expense_ids])
        where_clause = f"id IN ({ids_str}) AND (synced_to_valor = FALSE OR synced_to_valor IS NULL)"
    else:
        where_clause = "(synced_to_valor = FALSE OR synced_to_valor IS NULL)"
    
    query = f"""
    SELECT id, date, description, card_member, amount, category, project
    FROM `{PROJECT_ID}.{DATASET_ID}.{TABLE_ID}`
    WHERE {where_clause}
    """
    
    try:
        results = list(get_bq_client().query(query).result())
        
        if not results:
            return {"success": True, "synced_count": 0, "message": "No expenses to sync"}
        
        # Use MERGE to upsert - prevents duplicates!
        batch_size = 100
        total_synced = 0
        
        for i in range(0, len(results), batch_size):
            batch = results[i:i+batch_size]
            
            # Build VALUES for the merge
            values_parts = []
            for row in batch:
                valor_id = f"michael_{row.id}"  # Prefix with michael_ for unique ID
                date_str = row.date.isoformat() if row.date else None
                amount = float(row.amount) if row.amount else 0
                year = row.date.year if row.date else datetime.now().year
                month = row.date.month if row.date else datetime.now().month
                # Escape single quotes in strings
                name = (row.card_member or "Michael Nicklas").replace("'", "\\'")
                category = (row.category or "").replace("'", "\\'")
                project = (row.project or "").replace("'", "\\'")
                vendor = (row.description or "").replace("'", "\\'")
                
                values_parts.append(f"""
                    ('{valor_id}', '{name}', {amount}, '{category}', '{date_str}', '{vendor}', {year}, {month}, 'Michael Card', '{project}')
                """)
            
            values_sql = ",".join(values_parts)
            
            merge_query = f"""
                MERGE `{PROJECT_ID}.{DATASET_ID}.{VALOR_TABLE_ID}` AS target
                USING (
                    SELECT * FROM UNNEST([
                        STRUCT<id STRING, name STRING, amount FLOAT64, category STRING, date STRING, vendor STRING, year INT64, month INT64, source STRING, project STRING>
                        {values_sql}
                    ])
                ) AS source
                ON target.id = source.id
                WHEN MATCHED THEN
                    UPDATE SET
                        name = source.name,
                        amount = source.amount,
                        category = source.category,
                        date = PARSE_DATE('%Y-%m-%d', source.date),
                        vendor = source.vendor,
                        year = source.year,
                        month = source.month,
                        source = source.source,
                        project = source.project
                WHEN NOT MATCHED THEN
                    INSERT (id, created_at, name, amount, category, date, vendor, year, month, source, project)
                    VALUES (source.id, CURRENT_TIMESTAMP(), source.name, source.amount, source.category, 
                            PARSE_DATE('%Y-%m-%d', source.date), source.vendor, source.year, source.month, source.source, source.project)
            """
            
            get_bq_client().query(merge_query).result()
            total_synced += len(batch)
        
        # Mark all as synced in michael_expenses using UPDATE
        synced_ids = [row.id for row in results]
        batch_size = 500
        for i in range(0, len(synced_ids), batch_size):
            batch = synced_ids[i:i+batch_size]
            ids_str = ", ".join(f"'{id}'" for id in batch)
            
            update_query = f"""
                UPDATE `{PROJECT_ID}.{DATASET_ID}.{TABLE_ID}`
                SET synced_to_valor = TRUE
                WHERE id IN ({ids_str})
            """
            get_bq_client().query(update_query).result()
        
        return {
            "success": True,
            "synced_count": total_synced,
            "message": f"Synced {total_synced} expenses to valor_expenses (using MERGE - no duplicates)"
        }
        
    except Exception as e:
        print(f"Error syncing michael to valor: {e}")
        raise e


def get_michael_summary(year: int = None) -> Dict[str, Any]:
    """Get summary statistics for Michael expenses."""
    
    if year is None:
        year = datetime.now().year
    
    query = f"""
    SELECT 
        COUNT(*) as total_count,
        SUM(amount) as total_amount,
        SUM(CASE WHEN synced_to_valor = TRUE THEN 1 ELSE 0 END) as synced_count,
        SUM(CASE WHEN synced_to_valor = TRUE THEN amount ELSE 0 END) as synced_amount
    FROM `{PROJECT_ID}.{DATASET_ID}.{TABLE_ID}`
    WHERE EXTRACT(YEAR FROM date) = {year}
    """
    
    try:
        result = list(get_bq_client().query(query).result())[0]
        
        return {
            "total_count": result.total_count or 0,
            "total_amount": float(result.total_amount) if result.total_amount else 0,
            "synced_count": result.synced_count or 0,
            "synced_amount": float(result.synced_amount) if result.synced_amount else 0,
            "unsynced_count": (result.total_count or 0) - (result.synced_count or 0),
            "unsynced_amount": (float(result.total_amount) if result.total_amount else 0) - (float(result.synced_amount) if result.synced_amount else 0)
        }
        
    except Exception as e:
        print(f"Error getting michael summary: {e}")
        return {
            "total_count": 0,
            "total_amount": 0,
            "synced_count": 0,
            "synced_amount": 0,
            "unsynced_count": 0,
            "unsynced_amount": 0
        }

