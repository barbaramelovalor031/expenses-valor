"""
Rippling expense report processor
"""
import pandas as pd
from io import BytesIO
from typing import Dict, List, Any, Tuple

# Mapeamento de nomes do Rippling para (displayName, type)
EMPLOYEE_DATA: Dict[str, Tuple[str, str]] = {
    "Ana Coutinho": ("Ana Coutinho", "Contractor"),
    "Antoine Colaco": ("Antoine Colaco", "Partner"),
    "Barbara Melo": ("Barbara Melo", "Contractor"),
    "BARBARA MELO ADVISORY": ("Barbara Melo", "Contractor"),
    "Bernardo Rocha": ("Bernardo Rocha", "Contractor"),
    "BGFR CONSULTING BUSINESS INTELLIGENCE LTDA": ("Bernardo Rocha", "Contractor"),
    "Annelise Barre": ("Annelise Barre", "Contractor"),
    "Barre Negócios LTDA": ("Annelise Barre", "Contractor"),
    "Beatriz Balbuena": ("Beatriz Balbuena", "Contractor"),
    "Clifford Sobel": ("Clifford Sobel", "Partner"),
    "Bluebridge Advisers LLC": ("Clifford Sobel", "Partner"),
    "Bruno Batavia": ("Bruno Batavia", "Contractor"),
    "Pedro Bustamante": ("Pedro Bustamante", "Contractor"),
    "Bustamante LTDA": ("Pedro Bustamante", "Contractor"),
    "Caitlyn Oshman": ("Caitlyn Oshman", "Employee"),
    "Carlos Costa": ("Carlos Costa", "Partner"),
    "Carolina Hibner": ("Carolina Hibner", "Contractor"),
    "Carolina Ades Hibner LTDA": ("Carolina Hibner", "Contractor"),
    "DOUG SMITH": ("Doug Smith", "Partner"),
    "Doug Smith": ("Doug Smith", "Partner"),
    "Daniel Schulman": ("Daniel Schulman", "Partner"),
    "Fabiana Scionti": ("Fabiana Scionti", "Employee"),
    "Felipe Mendes": ("Felipe Mendes", "Contractor"),
    "Frances Townsend": ("Frances Townsend", "Advisor"),
    "Gabriel Gil": ("Gabriel Gil", "Contractor"),
    "GARGIL INTERMEDIACAO E AGENCIAMENTO LTDA": ("Gabriel Gil", "Contractor"),
    "Gustavo Nolla": ("Gustavo Nolla", "Contractor"),
    "GUSTAVO NOLLA CONSULTORIA LTDA": ("Gustavo Nolla", "Contractor"),
    "Gustavo Berger": ("Gustavo Berger", "Contractor"),
    "Karina Martinez": ("Karina Martinez", "Contractor"),
    "Kelli Spangler-Ballard": ("Kelli Spangler-Ballard", "Employee"),
    "Lana Brandao": ("Lana Brandao", "Contractor"),
    "Laura Pettinelli": ("Laura Pettinelli", "Employee"),
    "Marc Luongo": ("Marc Luongo", "Employee"),
    "Mario Mello": ("Mario Mello", "Partner"),
    "Michael Nicklas": ("Michael Nicklas", "Partner"),
    "Nicole Salim": ("Nicole Salim", "Contractor"),
    "NICOLE SALIM LTDA": ("Nicole Salim", "Contractor"),
    "Nicolas Marin": ("Nicolas Marin", "Contractor"),
    "Jose Noblecilla": ("Jose Noblecilla", "Contractor"),
    "PTECH": ("Jose Noblecilla", "Contractor"),
    "Paula Favaro": ("Paula Favaro", "Contractor"),
    "Paula Falcao Dufech Favaro LTDA": ("Paula Favaro", "Contractor"),
    "Paula Parnes": ("Paula Parnes", "Contractor"),
    "Paulo Passoni": ("Paulo Passoni", "Partner"),
    "Ricardo Villela Marino": ("Ricardo Villela Marino", "Partner"),
    "Scott Sobel": ("Scott Sobel", "Partner"),
    "Vivian Consolo": ("Vivian Consolo", "Contractor"),
    "Vivian Consolo Assessoria Executiva Ltda": ("Vivian Consolo", "Contractor"),
}

# Mapeamento de categorias do Rippling para categorias de output
CATEGORY_MAPPING: Dict[str, str] = {
    "Airfare": "Airfare",
    "Lodging": "Lodging",
    "Ground Transportation - Local": "Ground Transportation",
    "Ground Transportation - Travel": "Ground Transportation",
    "Meals & Entertainment - Local": "Meals & Entertainment",
    "Meals & Entertainment - Travel": "Meals & Entertainment",
    "Rippling Wire Deduction": "Rippling Wire Deduction",
    "IT Subscriptions": "IT Subscriptions",
    "Computer Equipment": "Computer Equipment",
    "Office Supplies": "Office Supplies",
    "Training": "Training",
    "Telephone/Internet": "Telephone/Internet",
    "Delivery and Postage": "Delivery and Postage",
    "Travel Agent Fees": "Travel Agent Fees",
    "Conferences & Seminars": "Conferences & Seminars",
    "Miscellaneous": "Miscellaneous",
}

# Todas as categorias de output possíveis
OUTPUT_CATEGORIES = [
    "Airfare",
    "Lodging",
    "Ground Transportation",
    "Meals & Entertainment",
    "Rippling Wire Deduction",
    "IT Subscriptions",
    "Computer Equipment",
    "Office Supplies",
    "Training",
    "Telephone/Internet",
    "Delivery and Postage",
    "Travel Agent Fees",
    "Conferences & Seminars",
    "Miscellaneous",
]


def find_employee_data(name: str) -> Tuple[str, str]:
    """Encontra dados do funcionário pelo nome"""
    if not name:
        return (name, "Unknown")
    
    # Match direto
    if name in EMPLOYEE_DATA:
        return EMPLOYEE_DATA[name]
    
    # Match case-insensitive
    name_lower = name.lower().strip()
    for key, data in EMPLOYEE_DATA.items():
        if key.lower() == name_lower:
            return data
    
    return (name, "Unknown")


def map_category(category: str) -> str:
    """Mapeia categoria do Rippling para categoria de output"""
    if not category:
        return "Miscellaneous"
    
    # Match direto
    if category in CATEGORY_MAPPING:
        return CATEGORY_MAPPING[category]
    
    # Match parcial
    for key, value in CATEGORY_MAPPING.items():
        if key.lower() in category.lower():
            return value
    
    return "Miscellaneous"


def process_rippling_file(file_content: bytes, filename: str) -> Dict[str, Any]:
    """
    Processa arquivo Rippling (CSV ou XLSX) e retorna dados agregados por funcionário e categoria
    """
    # Ler arquivo
    if filename.endswith('.csv'):
        df = pd.read_csv(BytesIO(file_content))
    else:
        df = pd.read_excel(BytesIO(file_content))
    
    # Verificar colunas necessárias
    required_cols = ['Employee', 'Amount', 'Category name']
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Coluna obrigatória não encontrada: {col}")
    
    # Mapear nomes de funcionários e tipos
    def get_employee_info(name):
        display_name, emp_type = find_employee_data(name)
        return pd.Series([display_name, emp_type])
    
    df[['Employee Name', 'Employee Type']] = df['Employee'].apply(get_employee_info)
    
    # Mapear categorias
    df['Category'] = df['Category name'].apply(map_category)
    
    # Converter Amount para numérico
    df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce').fillna(0)
    
    # Obter todas as categorias únicas que estão nos dados
    categories_in_data = sorted(df['Category'].unique().tolist())
    
    # Criar pivot table
    pivot = df.pivot_table(
        index=['Employee Name', 'Employee Type'],
        columns='Category',
        values='Amount',
        aggfunc='sum',
        fill_value=0
    ).reset_index()
    
    # Calcular total por funcionário
    category_cols = [col for col in pivot.columns if col not in ['Employee Name', 'Employee Type']]
    pivot['Total'] = pivot[category_cols].sum(axis=1)
    
    # Ordenar por nome
    pivot = pivot.sort_values('Employee Name')
    
    # Converter para lista de dicionários para JSON
    records = pivot.to_dict('records')
    
    # Calcular totais por categoria
    totals = {'Employee Name': 'TOTAL', 'Employee Type': ''}
    for cat in category_cols:
        totals[cat] = float(pivot[cat].sum())
    totals['Total'] = float(pivot['Total'].sum())
    
    return {
        'records': records,
        'categories': categories_in_data,
        'totals': totals,
        'row_count': len(df),
        'employee_count': len(records)
    }


def export_rippling_to_excel(data: Dict[str, Any]) -> bytes:
    """
    Exporta dados processados do Rippling para Excel
    """
    records = data['records']
    totals = data['totals']
    categories = data['categories']
    
    # Criar DataFrame
    df = pd.DataFrame(records)
    
    # Reordenar colunas: Employee Name, Employee Type, categorias ordenadas, Total
    cols = ['Employee Name', 'Employee Type'] + sorted(categories) + ['Total']
    # Filtrar apenas colunas que existem
    cols = [c for c in cols if c in df.columns]
    df = df[cols]
    
    # Adicionar linha de totais
    totals_row = pd.DataFrame([totals])
    totals_row = totals_row[cols]
    df = pd.concat([df, totals_row], ignore_index=True)
    
    # Criar Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Rippling Report')
        
        # Formatar
        worksheet = writer.sheets['Rippling Report']
        
        # Ajustar largura das colunas
        for idx, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            ) + 2
            worksheet.column_dimensions[chr(64 + idx) if idx <= 26 else f'A{chr(64 + idx - 26)}'].width = min(max_length, 20)
        
        # Formatar valores como moeda (colunas de categoria e Total)
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Header style
        header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Total row style
        total_row = len(df) + 1
        total_fill = PatternFill(start_color='e6f2ff', end_color='e6f2ff', fill_type='solid')
        total_font = Font(bold=True)
        
        for cell in worksheet[total_row]:
            cell.fill = total_fill
            cell.font = total_font
        
        # Number format for currency columns
        for row in range(2, len(df) + 2):
            for col in range(3, len(cols) + 1):  # Começando da coluna 3 (categorias)
                cell = worksheet.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
    
    output.seek(0)
    return output.getvalue()
